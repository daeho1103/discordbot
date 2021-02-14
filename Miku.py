import discord
import asyncio
import random
import urllib.parse
import urllib
import youtube_dl
import openpyxl
import datetime
import time
from discord.utils import get
from discord.ext import commands

import os

access_token = os.environ["BOT_TOKEN"]
token = access_token

client = commands.Bot(command_prefix='!')
# intents = discord.Intents.default()
# client = discord.Client(intents=intents)

@client.event
async def on_ready():
    print("봇 준비 완료!")
    print(client.user.name)
    print("---------------")
    game = discord.Game("!명령어 라고 쳐주세요")
    await client.change_presence(status=discord.Status.online, activity=game)



@client.command()
async def 미쿠(ctx):
    await ctx.send("初音ミク 登場!")

@client.command()
async def 명령어(ctx):
    embed = discord.Embed(title="명령어", color=0x62c1cc)
    embed.set_thumbnail(url="https://search.pstatic.net/sunny/?src=https%3A%2F%2Fi.pinimg.com%2Foriginals%2F92%2Fc0%2F22%2F92c022cbbf748349b665390eff955579.jpg&type=sc960_832")
    embed.add_field(name="!미쿠",value="`하츠네 미쿠 등장!`")
    embed.add_field(name="!따라해",value="`따라하기`")
    embed.add_field(name="!입장",value="`음성 채팅 입장`")
    embed.add_field(name="!재생",value="`노래 재생 (재생 뒤 youtube url)`")
    embed.add_field(name="!퇴장",value="`음성 채팅 퇴장`")
    embed.add_field(name="!경고",value="`경고주기`")
    embed.add_field(name="!경고보기",value="`경고보기`")
    embed.add_field(name="!경고초기화",value="`경고초기화`")
    embed.add_field(name="!기능추가",value="`기능추가`")
    embed.add_field(name="!기능보기",value="`기능보기`")
    embed.add_field(name="!기능삭제",value="`기능삭제`")
    embed.add_field(name="!기능초기화",value="`기능초기화`")
    await ctx.send(embed=embed)

@client.command()
async def 따라해(ctx):
    msg = ctx.message.content[5:]
    await ctx.send(msg)


@client.command()
async def 입장(ctx):
    await ctx.message.author.voice.channel.connect()
    await ctx.message.channel.send("보이스채널 입장합니다.")
    
@client.command()
async def 퇴장(ctx):
    global voice
    for vc in client.voice_clients:
        if vc.guild == ctx.message.guild:
            voice = vc
        
    await voice.disconnect()
    await ctx.message.channel.send("보이스채널 퇴장합니다.")

@client.command()
async def 재생(ctx):
    global voice
    for vc in client.voice_clients:
        if vc.guild == ctx.message.guild:
            voice = vc

    url = ctx.message.content.split(" ")[1]
    option = {
        # 'outtmpl' : "file/" + url.split('=')[1] + ".mp3"
        'outtmpl' : "file/" + url.split('=')[1] + ".mp3",
        'format': 'bestaudio/best',
        'postprocessors': [{
            'key': 'FFmpegExtractAudio',
            'preferredcodec': 'mp3',
            'preferredquality': '192',
        }]
    }

    with youtube_dl.YoutubeDL(option) as ydl:
        ydl.download([url])
        info = ydl.extract_info(url, download=False)
        title = info["title"]

    voice.play(discord.FFmpegPCMAudio("file/" + url.split('=')[1] + ".mp3"))
    await ctx.message.channel.send(title + "을 재생합니다.")





@client.command()
async def 경고(ctx):
    msg = ctx.message.content[4:]
    file = openpyxl.load_workbook("경고.xlsx")
    sheet = file.active
    i = 1
    while True:
        if sheet["A" + str(i)].value == str(msg):
            sheet["B" + str(i)].value = int(sheet["B" + str(i)].value) + 1
            file.save("경고.xlsx")
            await ctx.message.channel.send("경고를 1회 받았습니다.")
            if sheet["B" + str(i)].value == 2:
                await ctx.message.channel.send("누적 2회")
            elif sheet["B" + str(i)].value == 3:
                await ctx.message.channel.send("누적 3회")
            elif sheet["B" + str(i)].value == 4:
                await ctx.message.channel.send("누적 4회")
            elif sheet["B" + str(i)].value == 5:
                await ctx.message.channel.send("누적 5회")
            elif sheet["B" + str(i)].value == 6:
                await ctx.message.channel.send("누적 6회")
            elif sheet["B" + str(i)].value == 7:
                await ctx.message.channel.send("누적 7회")
            elif sheet["B" + str(i)].value == 8:
                await ctx.message.channel.send("누적 8회")
            elif sheet["B" + str(i)].value == 9:
                await ctx.message.channel.send("누적 9회")
            elif sheet["B" + str(i)].value == 10:
                await ctx.message.channel.send("누적 10회 퇴장조치")
            break
        if sheet["A" + str(i)].value == None:
            sheet["A" + str(i)].value = str(msg)
            sheet["B" + str(i)].value = 1
            file.save("경고.xlsx")
            await ctx.message.channel.send("경고를 1회 받았습니다.")
            await ctx.message.channel.send("누적 1회")
            break
        i+=1

@client.command()
async def 경고보기(ctx):
    file = openpyxl.load_workbook("경고.xlsx")
    sheet = file.active
    i=1
    try:
        for i in range(1,100):
            await ctx.message.channel.send(sheet.cell(i,1).value)
            await ctx.message.channel.send(sheet.cell(i,2).value)
            i+=1
    except :
        await ctx.message.channel.send("--------")
        await ctx.message.channel.send("여기까지")

@client.command()
async def 경고초기화(ctx):
    file = openpyxl.load_workbook("경고.xlsx")
    sheet  = file.active
    sheet.delete_cols(1)
    sheet.delete_cols(1)
    file.save("경고.xlsx")
    await ctx.message.channel.send("경고가 초기화 되었습니다.")



@client.command()
async def 기능추가(ctx):
    msg = ctx.message.content[6:]
    file = openpyxl.load_workbook("기능추가.xlsx")
    sheet  = file.active
    i=1
    while True:
        if sheet["A" + str(i)].value == None:
            sheet["A" + str(i)].value = str(msg)
            file.save("기능추가.xlsx")
            await ctx.message.channel.send(msg + " 기능이 추가 되었습니다.")
            break
        i+=1

@client.command()
async def 기능보기(ctx):
    file = openpyxl.load_workbook("기능추가.xlsx")
    sheet = file.active
    i=1
    try:
        for i in range(1,100):
            await ctx.message.channel.send(sheet.cell(i,1).value)
            i+=1
    except :
        await ctx.message.channel.send("--------")
        await ctx.message.channel.send("여기까지")

@client.command()
async def 기능삭제(ctx):
    msg = ctx.message.content[6:]
    file = openpyxl.load_workbook("기능추가.xlsx")
    sheet  = file.active
    i=1
    while True:
        if sheet["A" + str(i)].value == str(msg):
            sheet.delete_rows(i)
            file.save("기능추가.xlsx")
            await ctx.message.channel.send(msg + " 기능이 삭제 되었습니다.")
            break
        i+=1

@client.command()
async def 기능초기화(ctx):
    file = openpyxl.load_workbook("기능추가.xlsx")
    sheet  = file.active
    sheet.delete_cols(1)
    file.save("기능추가.xlsx")
    await ctx.message.channel.send("기능이 초기화 되었습니다.")

@client.command()
async def 타이머(ctx):
    msg = ctx.message.content[5:]
    sec = int(msg)

    for i in range(sec, 0, -1):
        time.sleep(1)
        await ctx.message.channel.send(embed=discord.Embed(description= str(i) + "초"))
    else:
        await ctx.message.channel.send(embed=discord.Embed(description='타이머 종료'))

@client.command()
async def 알람추가(ctx):
    msg = ctx.message.content[6:]
    file = openpyxl.load_workbook("알람.xlsx")
    sheet  = file.active
    i=1
    while True:
        if sheet["A" + str(i)].value == None:
            li = msg.split(" ")
            da1 = li.pop(0)
            da2 = li.pop(0)
            da3 = li.pop(0)
            da4 = li.pop(0)
            da5 = li.pop(0)
            dd = da1 + da2 + da3 + da4 +da5 
            sheet["A" + str(i)].value = str(dd)
            file.save("알람.xlsx")
            year = dd[:4]
            month = dd[4:6]
            day = dd[6:8]
            hour = dd[8:10]
            minu = dd[10:]
            await ctx.message.channel.send(year + "년 " + month + '월 ' + day + "일 " + hour + '시 ' + minu + "분 알람이 추가되었습니다.")
            break
        i+=1

@client.command()
async def 알람삭제(ctx):
    msg = ctx.message.content[6:]
    file = openpyxl.load_workbook("알람.xlsx")
    sheet  = file.active
    i=1
    while True:
        li = msg.split(" ")
        da1 = li.pop(0)
        da2 = li.pop(0)
        da3 = li.pop(0)
        da4 = li.pop(0)
        da5 = li.pop(0)
        dd = da1 + da2 + da3 + da4 +da5 
        if sheet["A" + str(i)].value == str(dd):
            sheet.delete_rows(i)
            file.save("알람.xlsx")
            year = dd[:4]
            month = dd[4:6]
            day = dd[6:8]
            hour = dd[8:10]
            minu = dd[10:]
            await ctx.message.channel.send(year + "년 " + month + '월 ' + day + "일 " + hour + '시 ' + minu + "분 알람이 삭제 되었습니다.")
            break
        i+=1

@client.command()
async def 알람보기(ctx):
    file = openpyxl.load_workbook("알람.xlsx")
    sheet  = file.active
    i=1
    try:
        for i in range(1,100):
            dd = sheet.cell(i,1).value
            year = dd[:4]
            month = dd[4:6]
            day = dd[6:8]
            hour = dd[8:10]
            minu = dd[10:]
            await ctx.message.channel.send(year + "년 " + month + '월 ' + day + "일 " + hour + '시 ' + minu + "분")
            i+=1
    except :
        await ctx.message.channel.send("--------")
        await ctx.message.channel.send("여기까지")

@client.command()
async def 알람초기화(ctx):
    file = openpyxl.load_workbook("알람.xlsx")
    sheet  = file.active
    sheet.delete_cols(1)
    file.save("알람.xlsx")
    await ctx.message.channel.send("알람이 초기화 되었습니다.")
    

client.run(token)
